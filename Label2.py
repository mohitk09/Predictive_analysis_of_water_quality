import xlrd
from openpyxl import Workbook

wb = xlrd.open_workbook('Parameters.xlsx')
sh = wb.sheet_by_name('Sheet1')
wb = Workbook()
worksheet = wb.active

training_examples = []
initial_training_set = []
num_of_rows = sh.nrows
num_of_cols = sh.ncols

for i in range(4,15):
    worksheet.cell(row=1, column=i-2).value = sh.cell_value(i-1,1)
for i in range(4,15):
    worksheet.cell(row=2, column=i-2).value = sh.cell_value(i-1,2)
worksheet.cell(row=2,column=13).value = "Label"

limitsu = [5,15,18,8.5,800,200,200,250,0.3,500,5]
limitsl = [3,0,0,6.5,0,0,0,0,0,0,1]

for i in range(3,14):
    for j in range(3,6):
        training_examples.append(sh.cell_value(i,j))
    initial_training_set.append(training_examples)
    training_examples = []

print(initial_training_set)
sol = [-1, -1, -1, -1, -1, -1, -1, -1, -1, -1, -1]
final_training_set = []


def pert(level, sol1):
    for i in range(0, 3):
        temp_training_examples = []
        sol1[level] = i
        if level+1 <= 10:
            pert(level+1,sol1)
        if level == 10:
            for j in range(0, 11):
                temp_training_examples.append(initial_training_set[j][sol1[j]])
            final_training_set.append(temp_training_examples)


pert(0,sol)
row =3
col = 2
total_pos_res = 0

for i in range(0,len(final_training_set)):
    pos = 0
    for j in range(0,len(final_training_set[0])):
        worksheet.cell(row=row, column=col).value = final_training_set[i][j]
        if(final_training_set[i][j]>=limitsl[pos] and final_training_set[i][j]<=limitsu[pos]):
            pos = pos+1
        col = col + 1
    if(pos==11):
        worksheet.cell(row = row,column= col).value = 1
        total_pos_res = total_pos_res+1
    else:
        worksheet.cell(row = row,column = col).value = 0
    row  = row+1
    print(i)
    col = 2

print("positive results are: ",total_pos_res)

wb.save('Parameters1.xlsx')